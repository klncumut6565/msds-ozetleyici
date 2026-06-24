#!/usr/bin/env python3
"""
MSDS / SDS Özetleyici — Çevrimdışı Python Uygulaması
Ollama (yerel AI) + Streamlit + pdfplumber
"""

import streamlit as st
import streamlit.components.v1 as components
import pdfplumber
import requests
import json
import re
import io
import time
import zipfile
import base64
from datetime import datetime

# Online AI motoru (Gemini) — kurulu değilse uygulama yine açılır,
# sadece "Online" seçeneği devre dışı kalır.
try:
    from google import genai
    from google.genai import types as genai_types
    GEMINI_SDK_OK = True
except ImportError:
    GEMINI_SDK_OK = False

# ── SAYFA AYARLARI ──────────────────────────────────────────────
st.set_page_config(
    page_title="MSDS / SDS Özetleyici",
    page_icon="⚗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── GHS SEMBOLLERI ─────────────────────────────────────────────
GHS_LABELS = {
    "GHS01": "Patlayıcı",      "GHS02": "Yanıcı",
    "GHS03": "Oksitleyici",    "GHS04": "Basınçlı gaz",
    "GHS05": "Aşındırıcı",    "GHS06": "Akut toksisite",
    "GHS07": "Tahriş edici",  "GHS08": "Sağlık tehlikesi",
    "GHS09": "Çevre tehlikesi"
}

GHS_SVG = {
    "GHS01": """<line x1="50" y1="55" x2="50" y2="8" stroke="#000" stroke-width="6" stroke-linecap="round"/>
        <line x1="50" y1="55" x2="82" y2="23" stroke="#000" stroke-width="6" stroke-linecap="round"/>
        <line x1="50" y1="55" x2="18" y2="23" stroke="#000" stroke-width="6" stroke-linecap="round"/>
        <line x1="50" y1="55" x2="88" y2="55" stroke="#000" stroke-width="6" stroke-linecap="round"/>
        <line x1="50" y1="55" x2="12" y2="55" stroke="#000" stroke-width="6" stroke-linecap="round"/>
        <line x1="50" y1="55" x2="72" y2="90" stroke="#000" stroke-width="5" stroke-linecap="round"/>
        <line x1="50" y1="55" x2="28" y2="90" stroke="#000" stroke-width="5" stroke-linecap="round"/>
        <circle cx="50" cy="62" r="22" fill="#000"/>
        <path d="M50,40 Q60,30 50,22 Q58,14 54,8" fill="none" stroke="#000" stroke-width="4" stroke-linecap="round"/>
        <circle cx="53" cy="7" r="5" fill="#000"/>""",
    "GHS02": """<path d="M50,82 C34,72 28,50 40,32 C42,42 46,46 48,42 C46,28 52,14 62,8 C56,24 60,36 54,42 C62,30 68,38 64,54 C68,44 70,50 64,66 C60,78 56,82 50,82 Z" fill="#000"/>""",
    "GHS03": """<ellipse cx="50" cy="74" rx="20" ry="12" fill="none" stroke="#000" stroke-width="6"/>
        <path d="M50,62 C38,52 34,36 44,24 C46,34 49,37 50,33 C48,20 55,10 63,6 C57,22 60,32 54,38 C62,26 68,34 64,50 C68,40 70,46 64,60 C60,68 54,64 50,62 Z" fill="#000"/>""",
    "GHS04": """<rect x="32" y="46" width="36" height="36" rx="5" fill="#000"/>
        <ellipse cx="50" cy="46" rx="18" ry="9" fill="#000"/>
        <rect x="45" y="26" width="10" height="16" rx="3" fill="#000"/>
        <rect x="30" y="24" width="40" height="7" rx="3.5" fill="#000"/>
        <rect x="28" y="79" width="44" height="7" rx="3.5" fill="#000"/>""",
    "GHS05": """<rect x="10" y="24" width="30" height="6" rx="2" fill="#000"/>
        <ellipse cx="22" cy="40" rx="4" ry="6" fill="#000"/>
        <ellipse cx="30" cy="50" rx="3" ry="5" fill="#000"/>
        <polyline points="10,56 16,67 20,60 26,72 32,63 40,70" fill="none" stroke="#000" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>
        <path d="M60,28 C68,23 80,26 80,40 C80,54 72,58 68,65 C65,72 63,78 58,80 L48,80 L46,52 L50,46 L52,28 Z" fill="#000"/>
        <ellipse cx="68" cy="48" rx="7" ry="8" fill="white"/>
        <ellipse cx="64" cy="12" rx="4" ry="6" fill="#000"/>
        <ellipse cx="73" cy="18" rx="3" ry="5" fill="#000"/>""",
    "GHS06": """<ellipse cx="50" cy="32" rx="20" ry="18" fill="#000"/>
        <ellipse cx="41" cy="30" rx="6" ry="7" fill="white"/>
        <ellipse cx="59" cy="30" rx="6" ry="7" fill="white"/>
        <rect x="47" y="39" width="6" height="5" fill="white"/>
        <rect x="36" y="48" width="28" height="9" rx="2" fill="#000"/>
        <rect x="38" y="50" width="5" height="8" fill="white"/>
        <rect x="45" y="50" width="5" height="8" fill="white"/>
        <rect x="52" y="50" width="5" height="8" fill="white"/>
        <line x1="26" y1="64" x2="74" y2="82" stroke="#000" stroke-width="8" stroke-linecap="round"/>
        <line x1="74" y1="64" x2="26" y2="82" stroke="#000" stroke-width="8" stroke-linecap="round"/>
        <circle cx="26" cy="64" r="6" fill="#000"/><circle cx="74" cy="64" r="6" fill="#000"/>
        <circle cx="26" cy="82" r="6" fill="#000"/><circle cx="74" cy="82" r="6" fill="#000"/>""",
    "GHS07": """<rect x="42" y="18" width="16" height="46" rx="8" fill="#000"/>
        <circle cx="50" cy="78" r="9" fill="#000"/>""",
    "GHS08": """<circle cx="50" cy="22" r="11" fill="#000"/>
        <path d="M42,34 L38,68 L46,68 L48,52 L52,52 L54,68 L62,68 L58,34 Z" fill="#000"/>
        <circle cx="50" cy="47" r="14" fill="#000"/>
        <line x1="50" y1="33" x2="50" y2="61" stroke="white" stroke-width="3.5" stroke-linecap="round"/>
        <line x1="36" y1="47" x2="64" y2="47" stroke="white" stroke-width="3.5" stroke-linecap="round"/>
        <line x1="40.5" y1="37.5" x2="59.5" y2="56.5" stroke="white" stroke-width="3.5" stroke-linecap="round"/>
        <line x1="59.5" y1="37.5" x2="40.5" y2="56.5" stroke="white" stroke-width="3.5" stroke-linecap="round"/>""",
    "GHS09": """<rect x="47" y="10" width="6" height="40" rx="2" fill="#000"/>
        <line x1="50" y1="20" x2="26" y2="10" stroke="#000" stroke-width="5" stroke-linecap="round"/>
        <line x1="50" y1="26" x2="74" y2="14" stroke="#000" stroke-width="5" stroke-linecap="round"/>
        <line x1="50" y1="16" x2="30" y2="6"  stroke="#000" stroke-width="4" stroke-linecap="round"/>
        <line x1="50" y1="32" x2="28" y2="22" stroke="#000" stroke-width="4" stroke-linecap="round"/>
        <line x1="50" y1="34" x2="74" y2="24" stroke="#000" stroke-width="4" stroke-linecap="round"/>
        <line x1="50" y1="14" x2="68" y2="6"  stroke="#000" stroke-width="4" stroke-linecap="round"/>
        <ellipse cx="52" cy="66" rx="19" ry="9" fill="#000"/>
        <path d="M33,66 L18,56 L18,76 Z" fill="#000"/>
        <circle cx="63" cy="63" r="4" fill="white"/>
        <line x1="60.5" y1="60.5" x2="65.5" y2="65.5" stroke="#000" stroke-width="1.5" stroke-linecap="round"/>
        <line x1="65.5" y1="60.5" x2="60.5" y2="65.5" stroke="#000" stroke-width="1.5" stroke-linecap="round"/>"""
}

# ── YARDIMCI FONKSİYONLAR ───────────────────────────────────────
def ok(v):
    return bool(v and str(v).strip() and str(v).strip() not in ("null", "None", "—", "-"))

def safe_filename(name: str, fallback: str = "MSDS_Ozet") -> str:
    """Ürün adını Windows dosya adına uygun, Türkçe karaktersiz hale getirir."""
    tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    s = (name or "").translate(tr_map)
    s = re.sub(r"[^\w\s-]", "", s).strip()
    s = re.sub(r"\s+", "_", s)
    return s or fallback

def make_diamond(code):
    if code not in GHS_SVG:
        return ""
    return f"""<div style="text-align:center;width:62px;">
  <svg viewBox="0 0 100 100" width="50" height="50" style="display:block;margin:0 auto;">
    <defs><clipPath id="dc{code}"><polygon points="50,3 97,50 50,97 3,50"/></clipPath></defs>
    <polygon points="50,3 97,50 50,97 3,50" fill="white" stroke="#cc0000" stroke-width="5"/>
    <g clip-path="url(#dc{code})">{GHS_SVG[code]}</g>
  </svg>
  <div style="font-size:8.5px;color:#cc0000;font-weight:500;margin-top:7px;line-height:1.2;">{GHS_LABELS[code]}</div>
</div>"""

def sh(icon, label, bg):
    return f'<div style="background:{bg};color:#fff;padding:4px 8px;font-size:9.5px;font-weight:500;">{icon} {label}</div>'

def pr(label, val):
    if not ok(val):
        return ""
    return f'<div style="display:flex;gap:4px;font-size:9.5px;border-bottom:.5px solid #f2f2f2;padding:1.5px 0;"><span style="color:#777;flex-shrink:0;min-width:95px;">{label}</span><span style="color:#111;font-weight:500;">{val}</span></div>'

def fab(icon, label, val):
    if not ok(val):
        return ""
    return f'<div style="background:#e3f2fd;border-radius:3px;padding:4px 6px;"><div style="font-size:9px;font-weight:500;color:#01579b;margin-bottom:1.5px;">{icon} {label}</div><div style="font-size:9px;color:#1a1a1a;line-height:1.5;">{val}</div></div>'

def kr(icon, label, val):
    if not ok(val):
        return ""
    return f'<div style="display:flex;gap:6px;padding:3px 0;border-bottom:.5px dotted #e0e0e0;"><div style="font-size:9px;font-weight:500;color:#1b5e20;min-width:105px;flex-shrink:0;">{icon} {label}</div><div style="font-size:9px;color:#333;">{val}</div></div>'


# ── HTML KARTI OLUŞTUR ──────────────────────────────────────────
def generate_html_card(s: dict, company: dict = None, fname: str = "") -> str:
    if company is None:
        company = {}

    ghs_list = [c for c in (s.get("ghs_piktogramlari") or []) if c in GHS_SVG]
    h_list   = s.get("h_ifadeleri") or []
    p_list   = s.get("p_ifadeleri") or []
    adr      = s.get("adr_bolum14") or {}
    fp       = s.get("fiziksel_ozellikler") or {}
    ia       = s.get("ilk_yardim") or {}
    sa       = s.get("saglik_tehlikeleri") or {}
    yn       = s.get("yangin") or {}
    kk       = s.get("kkd") or {}

    has_adr  = ok(adr.get("un_numarasi")) or ok(adr.get("sevkiyat_adi")) or ok(adr.get("adr_sinifi"))
    has_co   = ok(company.get("name")) or company.get("logo")
    theme    = company.get("color", "#1a237e")
    danger   = "TEH" in (s.get("sinyal_kelimesi") or "")
    etk      = adr.get("etiketler") or []
    pg_lbl   = {"I": " — Yüksek", "II": " — Orta", "III": " — Düşük"}

    # GHS elmasları
    ghs_html = "".join(make_diamond(c) for c in ghs_list) \
               or '<span style="font-size:10px;color:#bbb;">Belirlenmemiş</span>'

    # Filigran (antetli logo) — kontrast biraz artırıldı
    watermark = ""
    if company.get("logo"):
        watermark = f'<div style="position:absolute;inset:0;display:flex;align-items:center;justify-content:center;pointer-events:none;z-index:10;mix-blend-mode:multiply;"><img src="{company["logo"]}" alt="" style="width:34%;max-width:260px;object-fit:contain;opacity:.17;"></div>'

    # Kurumsal blok (kompakt — başlık satırının soluna girer)
    co_block = ""
    if has_co:
        logo_tag = f'<div style="width:32px;height:32px;border-radius:4px;background:#fff;display:flex;align-items:center;justify-content:center;overflow:hidden;flex-shrink:0;padding:2px;"><img src="{company["logo"]}" style="max-width:28px;max-height:28px;object-fit:contain;"></div>' if company.get("logo") else ""
        name_tag = f'<div style="font-size:11px;font-weight:500;line-height:1.2;">{company["name"]}</div>' if ok(company.get("name")) else ""
        dept_tag = f'<div style="font-size:8.5px;opacity:.6;line-height:1.2;">{company["dept"]}</div>' if ok(company.get("dept")) else ""
        txt = f'<div>{name_tag}{dept_tag}</div>' if (name_tag or dept_tag) else ""
        co_block = f'<div style="display:flex;align-items:center;gap:8px;padding-right:12px;margin-right:12px;border-right:1px solid rgba(255,255,255,0.2);flex-shrink:0;">{logo_tag}{txt}</div>'

    # Sinyal kelimesi
    signal_html = ""
    if ok(s.get("sinyal_kelimesi")):
        sig_bg = "#b71c1c" if danger else "#bf360c"
        signal_html = f'<div style="background:{sig_bg};padding:2px 10px;border-radius:2px;font-size:10px;font-weight:500;margin-bottom:5px;">⚠ {s["sinyal_kelimesi"]}</div>'

    # H+P ifadeleri
    h_html = ""
    if h_list:
        items = "".join(f'<div style="font-size:9px;color:#333;padding:1.5px 0;border-bottom:.5px dotted #eee;">{h}</div>' for h in h_list)
        h_html += f'<div style="font-size:9px;font-weight:500;color:#b71c1c;margin-bottom:2px;">TEHLİKE (H)</div>{items}'
    if p_list:
        items = "".join(f'<div style="font-size:9px;color:#333;padding:1.5px 0;">{p}</div>' for p in p_list[:5])
        h_html += f'<div style="font-size:9px;font-weight:500;color:#1565c0;margin-top:7px;margin-bottom:2px;">ÖNLEM (P)</div>{items}'

    # Sağlık etkileri
    saglik_rows = ""
    for lbl, field in [("Solunum","solunum"),("Deri","deri"),("Göz","goz"),("Yutma","yutma")]:
        if ok(sa.get(field)):
            saglik_rows += f'<div style="font-size:9px;color:#333;padding:1px 0;"><span style="color:#777;">{lbl}: </span>{sa[field]}</div>'
    saglik_html = f'<div style="margin-top:5px;padding-top:4px;border-top:.5px dotted #ddd;"><div style="font-size:9px;font-weight:500;color:#555;margin-bottom:2px;">Sağlık etkileri</div>{saglik_rows}</div>' if saglik_rows else ""

    # ADR Bölüm 14 — her zaman gösterilir (veri yoksa da bölüm görünür)
    kemler_top = adr.get("kemler_kodu") if ok(adr.get("kemler_kodu")) else "–"
    kemler_bot = (adr.get("un_numarasi") or "–").replace("UN", "").strip() or "–"
    etk_html   = "".join(f'<span style="background:#fff3e0;border:1px solid #e65100;color:#bf360c;font-size:8px;font-weight:700;padding:1px 5px;border-radius:2px;">{e}</span>' for e in etk) if etk else ""

    id_rows = ""
    for lbl, field, style in [
        ("BM Numarası", "un_numarasi", "color:#4a148c;font-weight:700;font-size:10px;"),
        ("ADR Sınıfı", "adr_sinifi", "color:#111;font-weight:500;"),
        ("Alt Tehlike", "alt_tehlike", "color:#111;font-weight:500;"),
    ]:
        val = adr.get(field) if ok(adr.get(field)) else "Belirtilmemiş"
        val_style = style if ok(adr.get(field)) else "color:#bbb;font-size:9px;"
        id_rows += f'<div style="display:flex;gap:4px;font-size:9.5px;padding:2px 0;"><span style="color:#888;min-width:82px;">{lbl}</span><span style="{val_style}">{val}</span></div>'
    if ok(adr.get("ambalaj_grubu")):
        pg = adr["ambalaj_grubu"]
        id_rows += f'<div style="display:flex;gap:4px;font-size:9.5px;padding:2px 0;"><span style="color:#888;min-width:82px;">Ambalaj Grubu</span><span style="color:#111;font-weight:500;">{pg}{pg_lbl.get(pg,"")}</span></div>'
    if ok(adr.get("tunel_kodu")):
        id_rows += f'<div style="display:flex;gap:4px;font-size:9.5px;padding:2px 0;align-items:center;"><span style="color:#888;min-width:82px;">Tünel Kodu</span><span style="background:#311b92;color:#fff;font-weight:700;font-size:9px;padding:1px 8px;border-radius:2px;">{adr["tunel_kodu"]}</span></div>'
    if ok(adr.get("kemler_kodu")):
        id_rows += f'<div style="display:flex;gap:4px;font-size:9.5px;padding:2px 0;"><span style="color:#888;min-width:82px;">Kemler Kodu</span><span style="color:#111;font-weight:500;">{adr["kemler_kodu"]}</span></div>'

    deniz_badge = f'<span style="font-size:8.5px;font-weight:500;padding:1px 8px;border-radius:2px;background:{"#b71c1c" if adr.get("deniz_kirletici")=="Evet" else "#1b5e20"};color:#fff;">{adr["deniz_kirletici"]}</span>' if ok(adr.get("deniz_kirletici")) else '<span style="font-size:9px;color:#bbb;">Belirtilmemiş</span>'
    sevkiyat = f'<div style="font-size:9.5px;color:#111;font-weight:500;line-height:1.5;margin-bottom:8px;">{adr["sevkiyat_adi"]}</div>' if ok(adr.get("sevkiyat_adi")) else '<div style="font-size:9px;color:#bbb;margin-bottom:8px;">Belirtilmemiş</div>'
    ozel     = f'<div style="font-size:9px;color:#333;line-height:1.6;">{adr["ozel_hukumler"]}</div>' if ok(adr.get("ozel_hukumler")) else '<span style="font-size:9px;color:#bbb;">Belirtilmemiş</span>'

    # Bu üründe taşıma verisi yoksa kullanıcıya net bilgi notu
    not_regulated = ""
    if not has_adr:
        not_regulated = '<div style="margin-bottom:6px;padding:4px 8px;background:#e8f5e9;border-radius:2px;border-left:3px solid #1b5e20;"><span style="font-size:8.5px;color:#1b5e20;line-height:1.4;">ℹ Belgede ADR/taşımacılık bilgisi belirtilmemiş. Ürün tehlikeli madde kapsamında olmayabilir; yine de orijinal belgenin Bölüm 14\'ünü kontrol edin.</span></div>'

    adr_section = f"""<div style="border-top:1px solid #e0e0e0;">
  {sh("🚛","Bölüm 14 — Taşımacılık Bilgileri (ADR / RID / IMDG)","#4a148c")}
  {not_regulated}
  <div style="display:grid;grid-template-columns:auto 1fr 1fr 1fr;">
    <div style="padding:8px 14px;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:6px;border-right:1px solid #e0e0e0;">
      <div style="width:62px;border:2.5px solid #e65100;border-radius:2px;overflow:hidden;font-family:monospace;text-align:center;box-shadow:1px 2px 4px rgba(0,0,0,.18);">
        <div style="background:#FF8F00;padding:4px;font-size:16px;font-weight:800;color:#1a1a1a;letter-spacing:1px;border-bottom:2px solid #e65100;">{kemler_top}</div>
        <div style="background:#FF8F00;padding:4px;font-size:14px;font-weight:800;color:#1a1a1a;">{kemler_bot}</div>
      </div>
      <div style="font-size:7.5px;color:#555;text-align:center;line-height:1.4;">Kemler<br>Plakası</div>
      {f'<div style="display:flex;gap:3px;flex-wrap:wrap;justify-content:center;">{etk_html}</div>' if etk_html else ""}
    </div>
    <div style="padding:8px 10px;border-right:1px solid #e0e0e0;">
      <div style="font-size:9px;font-weight:500;color:#4a148c;margin-bottom:4px;border-bottom:.5px solid #ede7f6;padding-bottom:2px;">Kimlik &amp; Sınıf</div>
      {id_rows}
    </div>
    <div style="padding:8px 10px;border-right:1px solid #e0e0e0;">
      <div style="font-size:9px;font-weight:500;color:#4a148c;margin-bottom:4px;border-bottom:.5px solid #ede7f6;padding-bottom:2px;">BM Sevkiyat Adı</div>
      {sevkiyat}
      <div style="display:flex;align-items:center;gap:6px;"><span style="font-size:9px;color:#666;">Deniz kirletici:</span>{deniz_badge}</div>
    </div>
    <div style="padding:8px 10px;">
      <div style="font-size:9px;font-weight:500;color:#4a148c;margin-bottom:4px;border-bottom:.5px solid #ede7f6;padding-bottom:2px;">Özel Hükümler</div>
      {ozel}
      <div style="margin-top:8px;padding:4px 6px;background:#ede7f6;border-radius:2px;"><span style="font-size:7.5px;color:#4a148c;line-height:1.4;">⚠ Taşımadan önce tam ADR belgesi ve sürücü talimatlarını kontrol edin.</span></div>
    </div>
  </div>
</div>"""

    # Footer
    co_footer = f'<span style="color:#888;font-weight:500;">{company.get("name","")}{" · " + company.get("dept","") if ok(company.get("dept")) else ""} · </span>' if ok(company.get("name")) else ""
    urun       = s.get("urun_adi") or "—"
    kimyasal   = s.get("kimyasal_adi") or ""
    cas        = s.get("cas_numarasi") or ""
    formul     = s.get("formul") or ""
    uretici    = s.get("uretici") or "—"
    acil       = s.get("acil_telefon") or ""
    tehlike    = s.get("tehlike_sinifi") or ""
    cas_line   = f"CAS: {cas}" + (f" · {formul}" if ok(formul) else "")

    # PDF kaydederken dosya adı bu <title>'dan gelir → ürün adını dosya-dostu yap
    base_name = urun if ok(urun) and urun != "—" else (fname.replace(".pdf", "") if fname else "MSDS_Ozet")
    safe_title = safe_filename(base_name)

    return f"""<!DOCTYPE html>
<html lang="tr"><head>
<meta charset="UTF-8">
<title>{safe_title}</title>
<style>
@media print {{.no-print{{display:none!important}} #card{{box-shadow:none!important;border:none!important}} body{{padding:0;background:#fff;}}}}
*{{box-sizing:border-box}} body{{font-family:system-ui,Arial,sans-serif;margin:0;padding:16px;background:#f0f2f5;}}
</style></head><body>
<div class="no-print" style="display:flex;gap:8px;justify-content:flex-end;max-width:920px;margin:0 auto 12px;">
  <button onclick="printAs()" style="padding:6px 16px;background:#1565c0;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px;">🖨️ Yazdır / PDF Kaydet</button>
</div>
<script>
function printAs(){{
  var prev = document.title;
  try {{ document.title = "{safe_title}"; }} catch(e){{}}
  window.print();
  setTimeout(function(){{ try {{ document.title = prev; }} catch(e){{}} }}, 800);
}}
</script>
<div id="card" style="position:relative;isolation:isolate;max-width:920px;margin:0 auto;background:#fff;border:.5px solid #c8c8c8;border-radius:4px;overflow:hidden;">
{watermark}
<div style="background:{theme};color:#fff;padding:8px 14px;">
  <div style="display:flex;justify-content:space-between;align-items:center;">
    <div style="display:flex;align-items:center;min-width:0;">
      {co_block}
      <div style="min-width:0;">
        {"" if has_co else '<div style="font-size:8px;opacity:.5;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:2px;">Malzeme Güvenlik Bilgi Formu · Tek Sayfa Özet</div>'}
        <div style="font-size:17px;font-weight:500;line-height:1.15;">{urun}</div>
        {f'<div style="font-size:9.5px;opacity:.65;margin-top:1px;">{kimyasal}</div>' if ok(kimyasal) and kimyasal!=urun else ""}
      </div>
    </div>
    <div style="text-align:right;flex-shrink:0;margin-left:16px;">
      {signal_html}
      {f'<div style="font-size:9.5px;opacity:.65;">{cas_line}</div>' if ok(cas) else ""}
      <div style="font-size:8.5px;opacity:.45;margin-top:2px;">{datetime.now().strftime("%d.%m.%Y")}</div>
    </div>
  </div>
</div>
  </div>
</div>
<div style="background:#e8eaf6;padding:4px 14px;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid #c5cae9;">
  <span style="font-size:10px;color:#333;"><span style="color:#666;">Üretici/tedarikçi: </span><span style="font-weight:500;">{uretici}</span></span>
  {f'<span style="font-size:10px;color:#b71c1c;font-weight:500;">📞 Acil: {acil}</span>' if ok(acil) else ""}
</div>
<div style="display:grid;grid-template-columns:190px 1fr 1fr;border-bottom:1px solid #e0e0e0;">
  <div style="border-right:1px solid #e0e0e0;">
    {sh("⚠","GHS Piktogramları","#b71c1c")}
    <div style="padding:10px;display:flex;flex-wrap:wrap;gap:10px;justify-content:center;">{ghs_html}</div>
    {f'<div style="padding:0 10px 8px;font-size:9px;color:#555;border-top:.5px dotted #e0e0e0;padding-top:5px;">{tehlike}</div>' if ok(tehlike) else ""}
  </div>
  <div style="border-right:1px solid #e0e0e0;">
    {sh("⚗","Fiziksel Özellikler","#4527a0")}
    <div style="padding:5px 10px;">
      {pr("Görünüm",fp.get("gorunum"))}{pr("Renk",fp.get("renk"))}{pr("Koku",fp.get("koku"))}
      {pr("Parlama noktası",fp.get("parlama_noktasi"))}{pr("Kaynama noktası",fp.get("kaynama_noktasi"))}
      {pr("Erime noktası",fp.get("erime_noktasi"))}{pr("Yoğunluk",fp.get("yogunluk"))}
      {pr("pH",fp.get("ph"))}{pr("Çözünürlük",fp.get("cozunurluk"))}{pr("Buhar basıncı",fp.get("buhar_basinci"))}
    </div>
  </div>
  <div>
    {sh("🔴","Tehlike &amp; Önlem İfadeleri","#c62828")}
    <div style="padding:5px 10px;">{h_html}</div>
  </div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr;border-bottom:1px solid #e0e0e0;">
  <div style="border-right:1px solid #e0e0e0;">
    {sh("❤","İlk Yardım","#0277bd")}
    <div style="padding:6px 10px;">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;">
        {fab("👁","Göz",ia.get("goz"))}{fab("🖐","Deri",ia.get("deri"))}
        {fab("💨","Solunum",ia.get("solunum"))}{fab("💊","Yutma",ia.get("yutma"))}
      </div>
      {saglik_html}
    </div>
  </div>
  <div>
    {sh("🛡","Kişisel Koruyucu Donanım (KKD)","#1b5e20")}
    <div style="padding:6px 10px;">
      {kr("💨","Solunum koruma",kk.get("solunum"))}{kr("🧤","El koruma",kk.get("el"))}
      {kr("🥽","Göz/yüz koruma",kk.get("goz"))}{kr("🥼","Vücut koruma",kk.get("vucut"))}
    </div>
  </div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr 1fr;border-bottom:1px solid #e0e0e0;">
  <div style="border-right:1px solid #e0e0e0;">
    {sh("🔥","Yangın Önlemleri","#bf360c")}
    <div style="padding:5px 10px;font-size:9px;">
      {f'<div style="margin-bottom:4px;"><div style="font-weight:500;color:#555;margin-bottom:1px;">Söndürücü</div><div style="color:#333;">{yn.get("sondurme_araci")}</div></div>' if ok(yn.get("sondurme_araci")) else ""}
      {f'<div style="margin-bottom:4px;"><div style="font-weight:500;color:#b71c1c;margin-bottom:1px;">✕ Yasak söndürücü</div><div style="color:#333;">{yn.get("yasakli_sondurme")}</div></div>' if ok(yn.get("yasakli_sondurme")) else ""}
      {f'<div><div style="font-weight:500;color:#555;margin-bottom:1px;">Özel tehlike</div><div style="color:#333;">{yn.get("ozel_tehlike")}</div></div>' if ok(yn.get("ozel_tehlike")) else ""}
    </div>
  </div>
  <div style="border-right:1px solid #e0e0e0;">
    {sh("📦","Depolama","#e65100")}
    <div style="padding:5px 10px;font-size:9px;color:#333;line-height:1.6;">{s.get("depolama") or "—"}</div>
  </div>
  <div>
    {sh("💧","Dökülme &amp; Bertaraf","#37474f")}
    <div style="padding:5px 10px;font-size:9px;">
      {f'<div style="margin-bottom:4px;"><div style="font-weight:500;color:#333;margin-bottom:1px;">Müdahale</div><div style="color:#444;line-height:1.5;">{s.get("dokulmede_yapilacaklar")}</div></div>' if ok(s.get("dokulmede_yapilacaklar")) else ""}
      {f'<div><div style="font-weight:500;color:#333;margin-bottom:1px;">Bertaraf</div><div style="color:#444;line-height:1.5;">{s.get("bertaraf")}</div></div>' if ok(s.get("bertaraf")) else ""}
    </div>
  </div>
</div>
{adr_section}
<div style="background:#fafafa;padding:5px 14px;border-top:.5px solid #e0e0e0;display:flex;justify-content:space-between;align-items:center;font-size:8px;color:#aaa;">
  <span>Bu özet bilgi amaçlıdır; orijinal MSDS/SDS belgesinin yerini tutmaz.</span>
  <span>{co_footer}<span>{fname} · Çevrimdışı AI</span></span>
</div>
</div></body></html>"""


# ── PDF METİN ÇIKAR ────────────────────────────────────────────
def extract_pdf_text(file_bytes: bytes) -> str:
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        pages = []
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                pages.append(f"[Sayfa {i+1}]\n{text}")
        return "\n\n".join(pages)


# ── OLLAMA API ─────────────────────────────────────────────────
OLLAMA_PROMPT = """Bu bir MSDS/SDS (Malzeme Güvenlik Bilgi Formu) belgesidir.
Aşağıdaki metni analiz et. SADECE geçerli bir JSON nesnesi döndür — başka hiçbir metin yazma.
Türkçe yanıt ver. Bilgi yoksa null kullan. Listeler boş dizi olabilir.

BELGE:
{text}

JSON ŞEMASI:
{{"urun_adi":"","kimyasal_adi":null,"cas_numarasi":null,"formul":null,"uretici":null,"acil_telefon":null,"sinyal_kelimesi":"TEHLİKE veya UYARI veya null","ghs_piktogramlari":["GHS01"..."GHS09"],"tehlike_sinifi":null,"h_ifadeleri":["H225 - Çok kolay alev alır" max 7],"p_ifadeleri":["P210 - ..." en önemli 5],"fiziksel_ozellikler":{{"gorunum":null,"renk":null,"koku":null,"parlama_noktasi":null,"kaynama_noktasi":null,"erime_noktasi":null,"yogunluk":null,"ph":null,"cozunurluk":null,"buhar_basinci":null}},"saglik_tehlikeleri":{{"solunum":null,"deri":null,"goz":null,"yutma":null}},"ilk_yardim":{{"solunum":null,"deri":null,"goz":null,"yutma":null}},"yangin":{{"sondurme_araci":null,"yasakli_sondurme":null,"ozel_tehlike":null}},"kkd":{{"solunum":null,"el":null,"goz":null,"vucut":null}},"depolama":null,"dokulmede_yapilacaklar":null,"bertaraf":null,"adr_bolum14":{{"un_numarasi":null,"kemler_kodu":null,"sevkiyat_adi":null,"adr_sinifi":null,"alt_tehlike":null,"ambalaj_grubu":null,"deniz_kirletici":"Evet/Hayır/null","tunel_kodu":null,"etiketler":[],"ozel_hukumler":null}}}}"""


def call_ollama(text: str, model: str, base_url: str) -> dict:
    payload = {
        "model": model,
        "prompt": OLLAMA_PROMPT.format(text=text[:7000]),
        "stream": False,
        "format": "json",
        "options": {"temperature": 0.1}
    }
    resp = requests.post(f"{base_url.rstrip('/')}/api/generate", json=payload, timeout=300)
    resp.raise_for_status()
    raw = resp.json().get("response", "{}")
    raw = re.sub(r"```json|```", "", raw).strip()
    return json.loads(raw)


def call_gemini(text: str, api_key: str, model: str = "gemini-2.5-flash", max_retries: int = 3) -> dict:
    """Google Gemini API ile analiz. Ücretsiz katmanda günlük/dakikalık limitler
    olduğundan 429 (rate limit) hatasında kısa bir bekleme ile otomatik tekrar dener."""
    if not GEMINI_SDK_OK:
        raise RuntimeError(
            "google-genai kütüphanesi kurulu değil. Komut isteminde şunu çalıştırın:\n"
            "pip install google-genai"
        )
    client = genai.Client(api_key=api_key)
    last_err = None
    for attempt in range(max_retries):
        try:
            resp = client.models.generate_content(
                model=model,
                contents=OLLAMA_PROMPT.format(text=text[:15000]),
                config=genai_types.GenerateContentConfig(
                    temperature=0.1,
                    response_mime_type="application/json",
                ),
            )
            raw = (resp.text or "{}").strip()
            raw = re.sub(r"```json|```", "", raw).strip()
            return json.loads(raw)
        except json.JSONDecodeError:
            raise
        except Exception as e:
            last_err = e
            msg = str(e)
            if "429" in msg or "RESOURCE_EXHAUSTED" in msg or "rate" in msg.lower():
                time.sleep(2 ** (attempt + 1))  # 2s, 4s, 8s
                continue
            raise
    raise last_err


def call_ai(text: str, engine: str, model: str, ollama_url: str = "", gemini_api_key: str = "") -> dict:
    """Seçili AI motoruna göre tek bir çağrı noktası."""
    if engine == "gemini":
        return call_gemini(text, gemini_api_key, model)
    return call_ollama(text, model, ollama_url)


def build_excel_summary(records: list) -> bytes:
    """Toplu işlem sonuçlarını tek bir Excel (xlsx) dosyasına özetler — bir PDF, bir satır."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "MSDS Özet"

    headers = ["Dosya", "Ürün Adı", "Kimyasal Adı", "CAS No", "Üretici",
               "Sinyal Kelimesi", "GHS Piktogramları", "Tehlike Sınıfı",
               "H İfadeleri", "UN Numarası", "ADR Sınıfı", "Ambalaj Grubu", "Durum"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    for rec in records:
        if rec.get("error"):
            ws.append([rec["filename"], "", "", "", "", "", "", "", "", "", "", "", f"HATA: {rec['error']}"])
            continue
        s = rec["data"]
        adr = s.get("adr_bolum14") or {}
        ws.append([
            rec["filename"],
            s.get("urun_adi") or "",
            s.get("kimyasal_adi") or "",
            s.get("cas_numarasi") or "",
            s.get("uretici") or "",
            s.get("sinyal_kelimesi") or "",
            ", ".join(s.get("ghs_piktogramlari") or []),
            s.get("tehlike_sinifi") or "",
            "; ".join(s.get("h_ifadeleri") or []),
            adr.get("un_numarasi") or "",
            adr.get("adr_sinifi") or "",
            adr.get("ambalaj_grubu") or "",
            "Tamam"
        ])

    widths = [24, 26, 20, 14, 18, 13, 18, 16, 32, 13, 11, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def check_ollama(base_url: str):
    try:
        r = requests.get(f"{base_url.rstrip('/')}/api/tags", timeout=3)
        if r.ok:
            return True, [m["name"] for m in r.json().get("models", [])]
    except Exception:
        pass
    return False, []


# ── ANA UYGULAMA ───────────────────────────────────────────────
def main():
    # ── SIDEBAR ──
    with st.sidebar:
        st.header("⚙️ Ayarlar")

        st.subheader("🤖 Yapay Zeka Motoru")

        # Yerel (Ollama) gerçekten erişilebilir mi? Bulutta erişilemez → varsayılanı Gemini yap.
        _local_ok, _ = check_ollama("http://localhost:11434")
        engine_options = ["🖥️ Yerel (Ollama) — ücretsiz, donanıma bağlı",
                          "☁️ Online (Gemini API) — ücretsiz, hızlı"]
        default_idx = 0 if _local_ok else 1
        if not _local_ok:
            st.caption("ℹ️ Yerel AI (Ollama) bu cihazda bulunamadı — Online (Gemini) önerilir.")
        engine_label = st.radio(
            "Motor seç", engine_options, index=default_idx,
            label_visibility="collapsed",
        )
        engine = "gemini" if engine_label.startswith("☁️") else "ollama"

        ollama_url, ollama_ok, model, gemini_api_key = "http://localhost:11434", False, "", ""

        if engine == "ollama":
            ollama_url = st.text_input("Ollama adresi", "http://localhost:11434")
            ollama_ok, models = check_ollama(ollama_url)

            if ollama_ok and models:
                st.success(f"✓ Ollama aktif — {len(models)} model mevcut")
                preferred = [m for m in models if any(x in m.lower()
                             for x in ["qwen", "mistral", "llama3", "phi", "gemma", "deepseek"])]
                model = st.selectbox("Model seç", preferred or models)
            elif ollama_ok:
                st.warning("⚠ Ollama çalışıyor ama model yüklü değil")
                model = st.text_input("Model adı", "mistral")
                st.code("ollama pull qwen2.5")
            else:
                st.error("✗ Ollama bulunamadı")
                model = st.text_input("Model adı", "mistral")
                with st.expander("📥 Kurulum"):
                    st.markdown("""
**1.** [ollama.ai](https://ollama.ai) → İndir ve kur
**2.** Terminalde:
```bash
ollama pull qwen2.5
ollama serve
```
**3.** Bu uygulamayı tekrar başlat""")
        else:
            if not GEMINI_SDK_OK:
                st.error("✗ google-genai kurulu değil")
                st.code("pip install google-genai")
            model = st.selectbox("Model", ["gemini-2.5-flash", "gemini-2.5-flash-lite"])

            st.markdown(
                '🔑 **Ücretsiz Gemini API anahtarı al** → '
                '[aistudio.google.com/apikey](https://aistudio.google.com/apikey)',
                unsafe_allow_html=False
            )
            gemini_api_key = st.text_input(
                "Gemini API anahtarı", type="password",
                placeholder="AIza... (buraya yapıştır)",
                help="Anahtarınız yalnızca bu oturumda kullanılır, hiçbir yere kaydedilmez."
            )

            with st.expander("❓ Anahtarı nasıl alırım? (30 saniye)"):
                st.markdown(
                    "1. [aistudio.google.com/apikey](https://aistudio.google.com/apikey) adresine git\n"
                    "2. Google hesabınla giriş yap (kredi kartı gerekmez)\n"
                    "3. **\"Create API key\"** butonuna tıkla\n"
                    "4. Oluşan `AIza...` ile başlayan anahtarı kopyala\n"
                    "5. Yukarıdaki kutuya yapıştır — hepsi bu!"
                )

            with st.expander("ℹ️ Ücretsiz kullanım sınırları ve gizlilik"):
                st.markdown(
                    "- Flash modelinde günlük ~1.500 istek (≈1.500 PDF/gün)\n"
                    "- Dakikada ~10-15 istek — uygulama otomatik aralık koyar\n"
                    "- **Gizlilik:** Ücretsiz katmanda gönderilen içerik Google tarafından "
                    "model eğitiminde kullanılabilir. Gizli/hassas MSDS verisi için "
                    "uygulamayı bilgisayarınızda yerel (Ollama) modunda çalıştırın.\n"
                    "- Güncel limitler: ai.google.dev/gemini-api/docs/rate-limits"
                )

            # Kullanıcı anahtar girmediyse ve sahibi ortak bir anahtarı Secrets'a koyduysa onu kullan (opsiyonel)
            if not gemini_api_key:
                try:
                    shared_key = st.secrets.get("GEMINI_API_KEY", "")
                except Exception:
                    shared_key = ""
                if shared_key:
                    gemini_api_key = shared_key
                    st.caption("ℹ️ Ortak anahtar kullanılıyor.")

            ollama_ok = bool(gemini_api_key) and GEMINI_SDK_OK

        st.divider()
        st.subheader("🏢 Kurumsal Şablon")
        co_name  = st.text_input("Firma adı",  placeholder="Kimya A.Ş.")
        co_dept  = st.text_input("Birim",       placeholder="İSG / Kalite")
        co_logo  = st.file_uploader("Logo (PNG/JPG/SVG)", type=["png", "jpg", "jpeg", "svg"])
        co_color = st.color_picker("Tema rengi", "#1a237e")

        st.divider()
        st.caption("⚗️ MSDS Özetleyici v1.1\nOllama + Gemini · Toplu işlem · Excel")

    company = {"name": co_name, "dept": co_dept, "color": co_color, "logo": None}
    if co_logo:
        raw_bytes = co_logo.read()
        b64 = base64.b64encode(raw_bytes).decode()
        company["logo"] = f"data:{co_logo.type};base64,{b64}"

    ready = ollama_ok  # her iki motor için de "hazır mı" bayrağı bu isimde tutuluyor

    # ── ANA ALAN ──
    st.title("⚗️ MSDS / SDS Özetleyici")
    st.caption("Türkçe · GHS + ADR Bölüm 14 · Yerel veya Online AI")

    tab_single, tab_batch = st.tabs(["🔍 Tekli Analiz", "📦 Toplu İşlem (Batch)"])

    # ════════════════════════════════════════════════════════════
    # TEKLİ ANALİZ
    # ════════════════════════════════════════════════════════════
    with tab_single:
        uploaded = st.file_uploader(
            "MSDS/SDS PDF dosyasını sürükleyin veya seçin",
            type="pdf", key="single_uploader",
            label_visibility="collapsed"
        )

        if uploaded:
            col_info, col_btn = st.columns([3, 1])
            with col_info:
                st.info(f"📄 **{uploaded.name}** — {uploaded.size // 1024} KB")
            with col_btn:
                analyze = st.button("🔍 Analiz Et", type="primary",
                                    disabled=not ready, use_container_width=True)

            if not ready:
                if engine == "ollama":
                    st.warning("Analiz için sol panelden Ollama'yı kurun ve başlatın.")
                else:
                    st.warning("Analiz için sol panelden Gemini API anahtarınızı girin.")

            if analyze and ready:
                with st.spinner("📄 PDF okunuyor..."):
                    file_bytes = uploaded.read()
                    try:
                        pdf_text = extract_pdf_text(file_bytes)
                        char_count = len(pdf_text)
                        if char_count < 100:
                            st.error("PDF'den metin çıkarılamadı (taranmış/görsel PDF olabilir).")
                            return
                        st.success(f"✓ {char_count:,} karakter okundu")
                    except Exception as e:
                        st.error(f"PDF okuma hatası: {e}")
                        return

                wait_msg = (f"🤖 {model} analiz ediyor… Bu 30–120 saniye sürebilir"
                            if engine == "ollama" else
                            f"☁️ {model} analiz ediyor… Bu genelde 5–10 saniye sürer")
                with st.spinner(wait_msg):
                    try:
                        result = call_ai(pdf_text, engine, model,
                                          ollama_url=ollama_url, gemini_api_key=gemini_api_key)
                    except requests.exceptions.Timeout:
                        st.error("⏱ Zaman aşımı. Daha küçük bir model deneyin (llama3.2).")
                        return
                    except json.JSONDecodeError:
                        st.error("Model JSON formatında yanıt vermedi. Tekrar deneyin.")
                        return
                    except Exception as e:
                        st.error(f"Analiz hatası: {e}")
                        return

                st.success("✓ Analiz tamamlandı!")

                html_card = generate_html_card(result, company, uploaded.name)
                components.html(html_card, height=1350, scrolling=True)

                dl_name = safe_filename(result.get("urun_adi") or uploaded.name.replace(".pdf", ""))

                st.divider()
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.download_button(
                        "📄 HTML İndir (Yazdırılabilir)",
                        data=html_card.encode("utf-8"),
                        file_name=f"{dl_name}.html",
                        mime="text/html", use_container_width=True
                    )
                with c2:
                    st.download_button(
                        "📋 JSON İndir (Ham Veri)",
                        data=json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8"),
                        file_name=f"{dl_name}.json",
                        mime="application/json", use_container_width=True
                    )
                with c3:
                    st.download_button(
                        "🖨️ Yazdır → PDF için HTML'i Aç",
                        data=html_card.encode("utf-8"),
                        file_name=f"{dl_name}.html",
                        mime="text/html", use_container_width=True
                    )
        else:
            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown("### 📥 Adım 1\nSol panelden AI motorunu seç ve PDF'i yükle")
            with col2:
                st.markdown("### 🤖 Adım 2\n'Analiz Et' butonuna bas")
            with col3:
                st.markdown("### 📄 Adım 3\nGHS piktogramları + ADR Bölüm 14 dahil özeti indir")

    # ════════════════════════════════════════════════════════════
    # TOPLU İŞLEM (BATCH)
    # ════════════════════════════════════════════════════════════
    with tab_batch:
        st.markdown(
            "Yüzlerce PDF'i tek seferde sıraya koyup işleyebilirsin. "
            "Bittiğinde tek bir **Excel özeti** ve istersen tüm **HTML/JSON kartlarının ZIP'i** olarak indirebilirsin."
        )
        if engine == "gemini":
            st.caption("☁️ Online motor seçili — istekler arasına otomatik küçük bir bekleme eklenir "
                       "(ücretsiz katman dakikalık limiti için).")
        else:
            st.caption("🖥️ Yerel motor seçili — yüzlerce dosyada toplam süre donanımınıza göre saatler sürebilir. "
                       "Daha hızlı sonuç için sol panelden Gemini'ye geçebilirsiniz.")

        batch_files = st.file_uploader(
            "MSDS/SDS PDF dosyalarını seçin (çoklu seçim)",
            type="pdf", accept_multiple_files=True, key="batch_uploader"
        )

        if batch_files:
            st.info(f"📦 **{len(batch_files)}** dosya seçildi.")
            start_batch = st.button("🚀 Toplu Analizi Başlat", type="primary",
                                     disabled=not ready, use_container_width=True)
            if not ready:
                if engine == "ollama":
                    st.warning("Toplu analiz için sol panelden Ollama'yı kurun ve başlatın.")
                else:
                    st.warning("Toplu analiz için sol panelden Gemini API anahtarınızı girin.")

            if start_batch and ready:
                progress = st.progress(0.0)
                status = st.empty()
                results = []
                total = len(batch_files)

                for idx, f in enumerate(batch_files):
                    status.write(f"İşleniyor: **{f.name}** ({idx + 1}/{total})")
                    record = {"filename": f.name}
                    try:
                        pdf_text = extract_pdf_text(f.read())
                        if len(pdf_text) < 100:
                            record["error"] = "Metin çıkarılamadı (taranmış/görsel PDF olabilir)"
                        else:
                            data = call_ai(pdf_text, engine, model,
                                            ollama_url=ollama_url, gemini_api_key=gemini_api_key)
                            record["data"] = data
                            record["html"] = generate_html_card(data, company, f.name)
                    except Exception as e:
                        record["error"] = str(e)

                    results.append(record)
                    progress.progress((idx + 1) / total)

                    if engine == "gemini" and idx < total - 1:
                        time.sleep(2)  # ücretsiz katman dakikalık limiti için tampon

                status.empty()
                ok_count = sum(1 for r in results if "error" not in r)
                err_count = total - ok_count
                if err_count:
                    st.warning(f"✓ {ok_count} başarılı, ✗ {err_count} hatalı (detaylar Excel'de 'Durum' sütununda)")
                else:
                    st.success(f"✓ Tüm {ok_count} dosya başarıyla analiz edildi!")

                st.session_state["batch_results"] = results

        if "batch_results" in st.session_state:
            results = st.session_state["batch_results"]
            st.divider()
            st.subheader("📊 Sonuçlar")

            table_rows = []
            for r in results:
                if "error" in r:
                    table_rows.append({"Dosya": r["filename"], "Ürün Adı": "—", "Durum": f"❌ {r['error']}"})
                else:
                    table_rows.append({
                        "Dosya": r["filename"],
                        "Ürün Adı": r["data"].get("urun_adi") or "—",
                        "Durum": "✅ Tamam"
                    })
            st.dataframe(table_rows, use_container_width=True, hide_index=True)

            c1, c2, c3 = st.columns(3)
            with c1:
                excel_bytes = build_excel_summary(results)
                st.download_button(
                    "📊 Excel Özeti İndir (Tüm Sonuçlar)",
                    data=excel_bytes,
                    file_name=f"msds_toplu_ozet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with c2:
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for r in results:
                        if "html" in r:
                            nm = safe_filename((r.get("data") or {}).get("urun_adi") or r["filename"].replace(".pdf", ""))
                            zf.writestr(f"{nm}.html", r["html"])
                st.download_button(
                    "📄 Tüm HTML Kartları (ZIP)",
                    data=zip_buf.getvalue(),
                    file_name="msds_html_kartlari.zip",
                    mime="application/zip",
                    use_container_width=True
                )
            with c3:
                zip_buf2 = io.BytesIO()
                with zipfile.ZipFile(zip_buf2, "w", zipfile.ZIP_DEFLATED) as zf:
                    for r in results:
                        if "data" in r:
                            nm = safe_filename(r["data"].get("urun_adi") or r["filename"].replace(".pdf", ""))
                            zf.writestr(
                                f"{nm}.json",
                                json.dumps(r["data"], ensure_ascii=False, indent=2)
                            )
                st.download_button(
                    "📋 Tüm JSON Verisi (ZIP)",
                    data=zip_buf2.getvalue(),
                    file_name="msds_json_verileri.zip",
                    mime="application/zip",
                    use_container_width=True
                )

            if st.button("🗑️ Sonuçları temizle"):
                del st.session_state["batch_results"]
                st.rerun()


if __name__ == "__main__":
    main()
